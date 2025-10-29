"""Excel workbook rendering utilities for the Foundation Matching Tool."""

from __future__ import annotations

from collections import defaultdict
from datetime import timezone
from pathlib import Path
from typing import Dict, Iterable, List, Sequence

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .models import MatchResult, RunContext

SUMMARY_HEADERS = [
    ("Run Timestamp", "timestamp"),
    ("Operator", "operator"),
    ("Run Label", "run_label"),
    ("Faculty Source", "faculty_source"),
    ("Foundation Source", "foundation_source"),
    ("Total Matches", "total_matches"),
    ("Warnings", "warnings"),
]

DATA_HEADERS_BASE = [
    "Faculty Name",
    "Faculty Career Stage",
    "Foundation Name",
    "Match Score (0-100)",
    "Matched Keywords",
    "Matched Keyword Count",
    "Why Matched",
    "Funding Tier",
    "Deadline",
    "Website",
    "Eligibility Notes",
    "Areas of Funding",
    "Average Grant Amount",
    "Career Stage Targeted",
    "Institution Preferences",
    "Row Reference",
]

DATA_HEADERS_WEIGHTED = [
    "Faculty Name",
    "Faculty Career Stage",
    "Foundation Name",
    "Match Score (0-100)",
    "Weighted Score (0-100)",
    "Matched Keywords",
    "Matched Keyword Count",
    "Why Matched",
    "Funding Tier",
    "Deadline",
    "Website",
    "Eligibility Notes",
    "Areas of Funding",
    "Average Grant Amount",
    "Career Stage Targeted",
    "Institution Preferences",
    "Row Reference",
]

DATA_COLUMN_WIDTHS_BASE = [
    22,
    18,
    28,
    14,
    28,
    18,
    40,
    14,
    14,
    30,
    26,
    26,
    18,
    24,
    26,
    28,
]

DATA_COLUMN_WIDTHS_WEIGHTED = [
    22,
    18,
    28,
    14,
    14,
    28,
    18,
    40,
    14,
    14,
    30,
    26,
    26,
    18,
    24,
    26,
    28,
]


def group_matches_by_division(
    matches: Iterable[MatchResult],
) -> Dict[str, List[MatchResult]]:
    """Group matches by faculty division."""

    grouped: Dict[str, List[MatchResult]] = defaultdict(list)
    for match in matches:
        grouped[match.faculty.division].append(match)

    # Sort each division's matches by faculty name, score desc, then foundation name
    for division, items in grouped.items():
        items.sort(
            key=lambda m: (
                m.faculty.name.lower(),
                -m.score,
                m.foundation.name.lower(),
            )
        )

    if not grouped:
        grouped["Neonatology"] = []

    return dict(grouped)


def build_workbook(
    matches_by_division: Dict[str, Sequence[MatchResult]],
    run_context: RunContext,
) -> Workbook:
    """Construct an Excel workbook containing summary and division sheets."""

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"

    _populate_summary_sheet(summary_sheet, matches_by_division, run_context)

    for division, matches in matches_by_division.items():
        sheet = workbook.create_sheet(title=division)
        _populate_division_sheet(sheet, matches, run_context.weighted_mode)

    if run_context.weighted_mode:
        _add_meta_sheet(workbook, run_context)

    return workbook


def write_workbook(workbook: Workbook, destination: Path) -> Path:
    """Persist workbook to disk, ensuring directories exist."""

    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    return destination


def _populate_summary_sheet(
    sheet: Worksheet,
    matches_by_division: Dict[str, Sequence[MatchResult]],
    run_context: RunContext,
) -> None:
    total_matches = sum(len(matches) for matches in matches_by_division.values())

    values = {
        "timestamp": run_context.timestamp.isoformat(),
        "operator": run_context.operator,
        "run_label": run_context.run_label or "—",
        "faculty_source": str(run_context.faculty_source),
        "foundation_source": str(run_context.foundation_source),
        "total_matches": total_matches,
        "warnings": _format_warnings(run_context.warnings),
    }

    for index, (label, key) in enumerate(SUMMARY_HEADERS, start=1):
        sheet[f"A{index}"] = label
        sheet[f"B{index}"] = values[key]

    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 80


def _populate_division_sheet(
    sheet: Worksheet,
    matches: Sequence[MatchResult],
    weighted_mode: bool,
) -> None:
    headers = DATA_HEADERS_WEIGHTED if weighted_mode else DATA_HEADERS_BASE
    widths = DATA_COLUMN_WIDTHS_WEIGHTED if weighted_mode else DATA_COLUMN_WIDTHS_BASE

    sheet.append(headers)

    for column_index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column_index)].width = width

    sheet.freeze_panes = "A2"

    for match in matches:
        row = [
            match.faculty.name,
            match.faculty.career_stage,
            match.foundation.name,
            match.keyword_score,
        ]
        if weighted_mode:
            row.append(
                match.weighted_score
                if match.weighted_score is not None
                else match.score
            )
        row.extend(
            [
                ", ".join(sorted(match.matched_keywords)),
                match.keyword_match_count,
                match.match_reason or "—",
                match.foundation.average_grant_amount,
                match.foundation.deadlines,
                match.foundation.website,
                match.foundation.institution_preferences or "—",
                match.foundation.areas_of_funding,
                match.foundation.average_grant_amount,
                match.foundation.career_stage_targeted,
                match.foundation.institution_preferences or "—",
                f"Faculty row {match.faculty.raw_row_index} / Foundation row {match.foundation.raw_row_index}",
            ]
        )
        sheet.append(row)


def _format_warnings(warnings: Sequence[str]) -> str:
    if not warnings:
        return "None"
    return "\n".join(f"- {warning}" for warning in warnings)


def _add_meta_sheet(workbook: Workbook, run_context: RunContext) -> None:
    sheet = workbook.create_sheet(title="Meta")

    timestamp = run_context.timestamp
    if timestamp.tzinfo is None:
        timestamp_utc = timestamp.replace(tzinfo=timezone.utc)
    else:
        timestamp_utc = timestamp.astimezone(timezone.utc)

    rows = [
        ("FMT Version", "v2.2-dev"),
        ("Weighted Mode", str(run_context.weighted_mode)),
        (
            "Sorted By",
            "Weighted Score (0-100)"
            if run_context.weighted_mode
            else "Match Score (0-100)",
        ),
        ("Generated At UTC", timestamp_utc.isoformat()),
        ("Output Path", str(run_context.output_path)),
    ]

    for row_index, (label, value) in enumerate(rows, start=1):
        sheet[f"A{row_index}"] = label
        sheet[f"B{row_index}"] = value

    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 80
